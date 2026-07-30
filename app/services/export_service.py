"""CSV / Excel / PDF generation for exports and printable applications."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

BRAND = colors.HexColor("#123D33")


def to_csv(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 fine


def to_xlsx(headers: list[str], rows: list[list], title: str = "Export") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="123D33")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max(
            [len(str(headers[col - 1]))]
            + [len(str(row[col - 1])) for row in rows[:200] if col - 1 < len(row)]
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 50)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(headers: list[str], rows: list[list], title: str = "Export") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=12 * mm, title=title,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9)
    head_style = ParagraphStyle("head", parent=styles["Normal"], fontSize=8,
                                leading=10, textColor=colors.white)
    data = [[Paragraph(str(h), head_style) for h in headers]]
    for row in rows[:2000]:
        data.append([Paragraph(str(v) if v is not None else "", cell_style) for v in row])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9c9c9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f2f6f4")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    doc.build([
        Paragraph(title, ParagraphStyle("t", parent=styles["Title"], textColor=BRAND)),
        Spacer(1, 6),
        table,
    ])
    return buf.getvalue()


def application_pdf(app_dict: dict, notes: list[dict]) -> bytes:
    """A printable single-application sheet (Part 4: Download / Print)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Application {app_dict.get('application_no', '')}",
    )
    styles = getSampleStyleSheet()
    label = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=9,
                           textColor=colors.HexColor("#5a6b64"))
    value = ParagraphStyle("val", parent=styles["Normal"], fontSize=10)

    story = [
        Paragraph("Brains College — Admission Application",
                  ParagraphStyle("h", parent=styles["Title"], textColor=BRAND)),
        Paragraph(f"Application No: <b>{app_dict.get('application_no', '')}</b> "
                  f"&nbsp;&nbsp; Submitted: {app_dict.get('submitted_at', '')}",
                  styles["Normal"]),
        Spacer(1, 10),
    ]

    fields = [
        ("Full name", "full_name"), ("Father's name", "father_name"),
        ("CNIC", "cnic"), ("Phone", "phone"), ("Email", "email"),
        ("Gender", "gender"), ("Date of birth", "date_of_birth"),
        ("City", "city"), ("Address", "address"),
        ("Department", "department"), ("Course applied", "course"),
        ("Campus", "campus"),
        ("Previous qualification", "previous_qualification"),
        ("Percentage", "percentage"),
        ("Application status", "application_status"),
        ("Payment status", "payment_status"),
        ("Admission status", "admission_status"),
        ("Documents", "documents"),
    ]
    rows = []
    for lbl, key in fields:
        v = app_dict.get(key)
        if isinstance(v, list):
            v = ", ".join(str(x) for x in v)
        rows.append([Paragraph(lbl, label), Paragraph(str(v or "—"), value)])
    table = Table(rows, colWidths=[52 * mm, None])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d7ded9")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f2f6f4")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)

    if notes:
        story += [Spacer(1, 12),
                  Paragraph("Administrator notes", styles["Heading3"])]
        for n in notes:
            story.append(Paragraph(
                f"<b>{n.get('admin_name', 'Admin')}</b> "
                f"({n.get('created_at', '')}): {n.get('note', '')}", value))
            story.append(Spacer(1, 4))

    doc.build(story)
    return buf.getvalue()
